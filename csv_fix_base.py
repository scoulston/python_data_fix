#!/usr/bin/env python
# -*- coding: utf-8 -*-


import pandas as pd
import csv
import random
import re
from time import process_time
from openpyxl import load_workbook
from openpyxl.styles import PatternFill, Border, Side, Alignment, Protection, Font, Color, colors, NamedStyle, numbers
from openpyxl.utils import FORMULAE
from openpyxl.styles.differential import DifferentialStyle
from openpyxl.formatting.rule import Rule
from openpyxl.utils.cell import get_column_letter

#function for ordering columns
def order(frame,var):
    if type(var) is str:
        var = [var] #let the command take a string or list
    varlist =[w for w in frame.columns if w not in var]
    frame = frame[var+varlist]
    return frame

#if I do a setup, part of the setup could be setting these default paths
default_base_file_path = 'C:/Users/user/fixed_data' #Just the base file path to send the fixed file to.
default_junk_file_path = 'C:/Users/user/junk' #Just the base junk file path to send the junk/intermediate files to.'

def twent_cha_fix(str1):
    str2 = ""
    if re.search("^DB-", str1):
        str2 = str1[3:]
    elif re.search("^DMS-", str1):
        str2 = str1[4:]
    if len(str2) > 20:
        str2 = str2[:21]
    return str2



#All "start" and "end" variables are to calculate the time it takes to do certain sections of code
#Takes a full path to a csv_file to fix, a base file path where you want your fixed csv to go, and a junk_file_path where you want the junk files to go.
def fix_csv(csv_file_path, fixed_file_name = 'fixed_data.xlsx', base_file_path=default_base_file_path, junk_file_path= default_junk_file_path):
    csv.register_dialect('star', delimiter='*', quoting=csv.QUOTE_NONE)

    start = process_time()
    data = pd.read_csv(csv_file_path)

    print("Fixing bad description...")
    monster_descriptions = data['Item Description']

    monster_descriptions.to_csv(junk_file_path + 'monster_descriptions.csv', index = False)

    end = process_time()
    csv_open_export_descrip_time = "Opening csv and exporting description: " + str(end - start)

    start = process_time()
    with open(junk_file_path + 'monster_descriptions.csv', newline='') as monster_descriptions_read_csv, \
            open(junk_file_path + 'monster_descriptions_fixed.csv', 'w', newline='') as monster_descriptions_write_csv:

        missing = [6, 7]  # 1-indexed positions of missing values
        missing.sort()  # enforce the increasing order
        reader = csv.reader(monster_descriptions_read_csv, delimiter='*', skipinitialspace=True)
        writer = csv.writer(monster_descriptions_write_csv, delimiter = ',' )
        next(reader)
        header = 'metal_type','idk1','idk2','OD_or_H','od_name','ID_or_W','id_name','Length','length_name','cut','company','idk3','idk4', 'idk5' # get first row (header)
        writer.writerow(header)  # write it back
        for row in reader:
            if len(row) < 12:
                 #row shorter than header -> insert empty strings
                #inserting changes indices so `missing` must be sorted
                for idx in missing:
                    row.insert(idx - 1, '')
            writer.writerow(row)
    end = process_time()
    fix_monster_descrip_time = "Time to fix monster description: " + str(end - start)

    start = process_time()
    print("Fixing the date...")
    data=data.rename(columns = {'Item Description':'metal_type*idk1*idk2*OD_or_H*od_name*ID_or_W*id_name*Length*length_name*cut*company*idk3*idk4*idk5', 'Customer': 'Company'})
    #converts Day of Date column to a datetime column (to easily extract month, day and year from the column)
    data['Day of Date'] = pd.to_datetime(data['Day of Date'])
    data['Day of Date'] = data['Day of Date'].dt.strftime(r'%m/%d/%Y')

    # Create new columns for month, day and year (user requested change back to regular format, so commented out)
    #data['Month'] = data['Day of Date'].dt.month
    #data['Day'] = data['Day of Date'].dt.day
    #data['YR'] = data['Day of Date'].dt.year

    end = process_time()
    fix_date_time = "Time to fix date: " + str(end - start)




    start = process_time()
    monster_data = pd.read_csv(junk_file_path + 'monster_descriptions_fixed.csv')   
    
    print("Applying Company name fix...")
    #applies the fix to company where it takes out DMS- or DB- and only keeps 20characters after that, and strips extra space if any
    data['Company'] = data['Company'].apply(twent_cha_fix)
    data['Company'] = data['Company'].str.strip()

    print("Fixing rearranged columns for wacky rows that dont fit...")
    for index, row in monster_data.iterrows():
        if monster_data.loc[index, 'Length'] == 'L':
            monster_data.loc[index, 'Length'] = monster_data.loc[index, 'od_name']
            monster_data.loc[index, 'idk3'] = monster_data.loc[index, 'idk1']
            monster_data.loc[index, 'idk1'], monster_data.loc[index, 'idk2'] = monster_data.loc[index, 'idk2'], monster_data.loc[index, 'idk1']
            monster_data.loc[index, 'metal_type'] = " "
            monster_data.loc[index, 'company'] = monster_data.loc[index, 'cut']
            monster_data.loc[index, 'cut'] = " "
            monster_data.loc[index, 'idk2'] = monster_data.loc[index, 'OD_or_H']
            monster_data.loc[index, 'OD_or_H'] = " "
            

    #merges the description and rest of the data back together
    print("Merging everything back together...")
    data = pd.concat([data, monster_data], axis = 1, sort=False)

    #gets rid of unwanted columns
    print("Deleting unneeded columns...")
    #del data['Day of Date']
    del data['metal_type*idk1*idk2*OD_or_H*od_name*ID_or_W*id_name*Length*length_name*cut*company*idk3*idk4*idk5']
    del data['od_name']
    del data['id_name']
    del data['length_name']
    del data['company']
    del data['idk4']
    del data['idk5']

    #renames columns to better names
    print("Optimizing column names...")
    data.rename(columns = {'metal_type' : 'Shape', 'idk1' : 'Cond', 'idk2': 'Grade', 'Length': 'Len', 'Day of Date': 'Date'}, inplace=True)

    data.to_csv(junk_file_path + 'intermed_fix.csv', index=False) #This is the point fixing everything else but before fixing the triplicates
    end = process_time()
    add_columns_from_monster_data_time = "Time to add columns from monster data: " + str(end - start)

    start = process_time()
    print("Fixing triplicate rows...")
    triple_data = pd.read_csv(junk_file_path + 'intermed_fix.csv')
    measure_dict = {
        'Pounds Shipped': [],
        'Shipment Dollars': [],
        'QTY': []
    }
    #Converts appropriate numbers into a dictionary for the appropriate measure from the Measure Names Column
    for index, row in triple_data.iterrows():
        for key, value in measure_dict.items():
            if triple_data.loc[index, 'Measure Names'] == key:
                value.append(triple_data.loc[index, 'Measure Values'])
    measure_df = pd.DataFrame(measure_dict)

    fixed_triple_df = triple_data[triple_data.index % 3 == 0].reset_index(drop=True)  # Selects every 3rd raw starting from 0
    fixed_triple_df.reset_index(drop=True)
    final_df = pd.concat([fixed_triple_df, measure_df], axis=1)
    final_df.drop(columns=['Measure Names', 'Measure Values'], inplace = True)
    fixed_triple_df.to_csv(junk_file_path + 'test.csv')
    measure_df.to_csv(junk_file_path + 'measuretest.csv')
    #renames the Shipment Dollars to Sales
    final_df=final_df.rename(columns = {'Shipment Dollars': 'Sales'})

    print("Rounding the sales and pounds shipped...")
    #rounds the sales and pounds 
    final_df['Sales'] = final_df['Sales'].apply(round)
    final_df['Pounds Shipped'] = final_df['Pounds Shipped'].apply(round)

    print("Ordering the columns...")
    #orders the columns
    final_df = order(final_df, ['Company', 'Salesperson', 'Date', 'QTY', 'OD_or_H', 'ID_or_W', 'Len', 'Grade', 'Cond', 'Shape', 'Pounds Shipped', 'Sales'])
    #final_df.to_csv(base_file_path + fixed_file_name, index= False)
    #New code to test to switch to excel format
    final_df.to_excel (base_file_path + fixed_file_name, index = None, header=True)
    end = process_time()
    fix_triplicate_time = "Time to fix triplicate rows: " + str(end - start)

    #excel manipulation
    print("Messing with Excel formatting...")
    start = process_time()
    workbook = load_workbook(filename=base_file_path + fixed_file_name)
    sheet = workbook.active
    #inserts 3 rows before header for other stuff
    sheet.insert_rows(idx=1, amount=3)
    #freezes header row
    sheet.freeze_panes = "A5"
    #saves the header row to work with formatting
    header_row = sheet[4]

    #saves a header style
    header = NamedStyle(name="header")
    header.font = Font(bold=True)
    header.border = Border(
        bottom=Side(border_style="thick"), 
        left=Side(border_style="thick"), 
        right=Side(border_style="thick"), 
        top=Side(border_style="thick"))

    header.alignment = Alignment(horizontal="center", vertical="top", wrap_text=True)

    #applies header style to header row
    for cell in header_row:
        cell.style = header

    #Saves thin border  
    side = Side(border_style="thin", color="000000")
    border = Border(left=side, right=side, top=side, bottom=side)

    maxrow = sheet.max_row

    print("Fixing borders...")
    for i in range(1, sheet.max_column + 1):
        col = get_column_letter(i)
        for i in range(5, maxrow + 1):
            sheet[f"{col}{i}"].border = border

    print("Fixing number and text formats...")

    for i in range(4, sheet.max_column + 1):
        col = get_column_letter(i)
        if col == 'C':
            continue
        for i in range(5, maxrow + 1):
            if col in ['A', 'B', 'D', 'I', 'J', 'M', 'N', 'O', 'P', 'Q']:
                sheet[f"{col}{i}"].number_format = numbers.BUILTIN_FORMATS[49]
            elif col in ['E', 'F', 'G']:
                sheet[f"{col}{i}"].number_format = '0.000' #numbers.BUILTIN_FORMATS[4]
            else:
                sheet[f"{col}{i}"].number_format = numbers.BUILTIN_FORMATS[3]
    #adds title
    sheet["A1"] = "Customer Sales Details"
    sheet["A1"].font = Font(bold=True)

    #adds subtotals above pounds shipped and sales
    sheet["K3"] = f"=SUBTOTAL(9,K5:K{maxrow})"
    sheet["K3"].fill = PatternFill(start_color="9BFF9C", end_color="9BFF9C", fill_type = "solid")
    sheet["K3"].border = border
    sheet["K3"].number_format = numbers.BUILTIN_FORMATS[3]

    sheet["L3"] = f"=SUBTOTAL(9,L5:L{maxrow})"
    sheet["L3"].fill = PatternFill(start_color="9BFF9C", end_color="9BFF9C", fill_type = "solid")
    sheet["L3"].border = border
    sheet["L3"].number_format = numbers.BUILTIN_FORMATS[3]

    sheet["J3"] = "Totals:"
    sheet["J3"].font = Font(bold=True)
    sheet["J3"].fill = PatternFill(start_color="9BFF9C", end_color="9BFF9C", fill_type = "solid")
    sheet["J3"].border = border

    #adds filter buttons to columns
    sheet.auto_filter.ref = f"A4:Q{maxrow}"



    #saves the workbook
    workbook.save(base_file_path + fixed_file_name)

    end = process_time()
    excel_manip_time = "Time to manipulate excel: " + str(end - start) + "\n"
    #printing calc for time to do each section (can be commented out until needed)
    print(csv_open_export_descrip_time)
    print(fix_monster_descrip_time)
    #print(fix_date_time)
    print(add_columns_from_monster_data_time)
    print(fix_triplicate_time)
    print(excel_manip_time)

def input_csv_fix():
    path = input("Enter the path of the csv file to be fixed: ")
    #path = quote_path[1:-1] #use this if the copied path has quotes around it already (copying a path from a windows 10 computer for instance)
    fixed_name = input("Enter what you want as the name of the fixed file: ") + ".xlsx"
    fix_csv(r"{path}".format(path = path), '{fixed_name}'.format(fixed_name = fixed_name))
    start = process_time()
    return fixed_name
