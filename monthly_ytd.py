from csv_fix_base import input_csv_fix, default_base_file_path
import pandas as pd
from datetime import date, datetime
from time import process_time
from openpyxl import load_workbook
from openpyxl.styles import PatternFill, Border, Side, Alignment, Protection, Font, Color, colors, NamedStyle, numbers
from openpyxl.utils import FORMULAE
from openpyxl.styles.differential import DifferentialStyle
from openpyxl.formatting.rule import Rule
from openpyxl.utils.cell import get_column_letter


#function that makes a csv file with summary of data (total sales and lbs shipped per month for each company, including a column for the salesman)
def input_monthly_ytd_summary(input_csv_fix=input_csv_fix):
    #runs input_csv_fix to fix a csv, then takes the fixed file name to open it and work with the fixed data to start summarizing
    fixed_file_name = input_csv_fix()
    #uses the fixed file name to open the fixed file
    print("Fixing the date...")
    summary_data = pd.read_excel(default_base_file_path + fixed_file_name, header=3)
    start = process_time()
    #takes the month and year from the data and strings them together into "YR Month" format
    summary_data['Date'] = pd.to_datetime(summary_data['Date'])
    summary_data['Month_YR'] = summary_data['Date'].dt.year.astype(str) + " " + summary_data['Date'].dt.month.astype(str)
    #makes a list of the month_YR column's unique values to use later in renaming meshed columns
    months = summary_data['Month_YR'].unique().tolist()
    #changes the list months to datetime so we can get the wanted formatting later when renaming meshed columns
    for i in range(len(months)):
        months[i] = datetime.strptime(months[i], '%Y %m')
    end = process_time()
    fix_date_as_datetime_time = "Time to fix date column and store as datetime: " + str(end - start)

    start = process_time()
    #groups the data by company, salesperson, and month_yr
    print("Grouping data by Company, Salesperson, and Date...")
    data_by_month_and_comp = summary_data.groupby(['Company', 'Salesperson', 'Month_YR']).agg(
    {
         'Sales':sum, # Sum Sales per group
         'Pounds Shipped': sum,  # Sum Pounds Shipped per group
            }

).reset_index()
    end = process_time()
    group_by_table_time = "Time to group by the table: " + str(end - start)


    #uncomment the below line to get a file (group_by_test.csv) to test the above grouped table
    #data_by_month_and_comp.to_csv(default_base_file_path + "group_by_test.csv")
    print("Creating multiindex table...")
    #found this function online that supports pivoting multiindex, multilevel tables (where duplicated indices are not a problem)
    def multiindex_pivot(df, index=None, columns=None, values=None):
        if index is None:
            names = list(df.index.names)
            df = df.reset_index()
        else:
            names = index
        list_index = df[names].values
        tuples_index = [tuple(i) for i in list_index] # hashable
        df = df.assign(tuples_index=tuples_index)
        df = df.pivot(index="tuples_index", columns=columns, values=values)
        tuples_index = df.index  # reduced
        index = pd.MultiIndex.from_tuples(tuples_index, names=names)
        df.index = index
        return df
    
    start = process_time()
    #uses the above function to make a multiindex, multilevel table with the company and salesperson as indexes, the month-yr as one level of columns, the sales and pounds shipped as another level of columns
    data_by_month_and_comp_pivot = multiindex_pivot(data_by_month_and_comp, index = ['Company', 'Salesperson'], columns= 'Month_YR', values = ['Sales', 'Pounds Shipped']).reset_index() 
    end = process_time()
    multiindex_table_time = "Time to make the multiindex table: " + str(end - start)

    print("Sorting columns by date (month and year)")
    start = process_time()
    #fixes number months to abbreviated months in column labels using the sorted months list and converting it to the datetime format we want while still keeping the columns in the right order
    data_by_month_and_comp_pivot.columns.set_levels([x.strftime('%b %Y') for x in sorted(months)], 1, inplace=True, verify_integrity = False)
    end = process_time()
    fix_months_column_headers_time = "Time to fix the months in column headers: " + str(end - start)

    start = process_time()
    print("Adding total sales, pounds, and avg price per pound columns...")
    #adds columns for the total sales and total lbs for that company/salesperson during the timeperiod the data runs across 
    data_by_month_and_comp_pivot.insert(2, 'Total Sales', data_by_month_and_comp_pivot['Sales'].sum(axis=1))
    data_by_month_and_comp_pivot.insert(3, 'Total Lbs', data_by_month_and_comp_pivot["Pounds Shipped"].sum(axis=1))
    data_by_month_and_comp_pivot.insert(4, 'Avg Price Per Lb', data_by_month_and_comp_pivot['Total Sales'] / data_by_month_and_comp_pivot['Total Lbs'])
    end = process_time()
    make_total_sales_and_lbs_time = "Time to make the total sales and lbs columns: " + str(end - start)

    print("Merging levels in table...")
    start = process_time()
    #fixes the stacked columns into one column name per column (meshes the column level names)
    data_by_month_and_comp_pivot.columns = list(map(" ".join, data_by_month_and_comp_pivot.columns))
    data_by_month_and_comp_pivot.reset_index()
    end = process_time()
    merge_levels_time = "Time to merge levels on columns: " + str(end - start)

    start = process_time()
    #saves the data to the fixed csv filename 
    data_by_month_and_comp_pivot.to_excel(default_base_file_path + fixed_file_name, index = None, header=True)
    end = process_time()
    ot_export_fixed_time = "Time to export as fixed file: " + str(end - start)

    #excel manipulation
    print("Messing with Excel formatting...")
    start = process_time()
    workbook = load_workbook(filename=default_base_file_path + fixed_file_name)
    sheet = workbook.active
    #inserts 3 rows before header for other stuff
    sheet.insert_rows(idx=1, amount=3)
    #freezes header row
    sheet.freeze_panes = "A5"
    #saves the header row to work with formatting
    header_row = sheet[4]

    #saves a header style
    header = NamedStyle(name="header")
    header.font = Font(bold=True)
    header.border = Border(
        bottom=Side(border_style="thick"), 
        left=Side(border_style="thick"), 
        right=Side(border_style="thick"), 
        top=Side(border_style="thick"))

    header.alignment = Alignment(horizontal="center", vertical="top", wrap_text=True)

    print("\n")
    #applies header style to header row
    for cell in header_row:
        cell.style = header

    #adds totals above months
    maxrow = sheet.max_row

    #saves the border settings thin and all around
    side = Side(border_style="thin", color="000000")
    border = Border(left=side, right=side, top=side, bottom=side)

    for i in range(6, sheet.max_column + 1):
        col = get_column_letter(i)
        cell = get_column_letter(i) + "3"
        sheet[cell] = f"=SUBTOTAL(9,{col}5:{col}{maxrow})"
        sheet[cell].fill = PatternFill(start_color="9BFF9C", end_color="9BFF9C", fill_type = "solid")
        sheet[cell].border = border
        sheet[cell].number_format = numbers.BUILTIN_FORMATS[3]
        
    #formats the cell before the totals cells 
    sheet["E3"] = "Totals:"
    sheet["E3"].font = Font(bold=True)
    sheet["E3"].fill = PatternFill(start_color="9BFF9C", end_color="9BFF9C", fill_type = "solid")
    sheet["E3"].border = border

    print("\n")

    sheet["A1"] = "Monthly Sales & Lbs"
    sheet["A1"].font = Font(bold=True)
    #Adds red background to avg prices that have avg price/lb > $6 
    red_background = PatternFill(start_color="FC6E5E", end_color="FC6E5E")
    diff_style = DifferentialStyle(fill=red_background)
    rule = Rule(type="expression", dxf=diff_style)
    rule.formula = ["$E5>6"]
    sheet.conditional_formatting.add(f"E5:E{maxrow}", rule)
    
    #fixes the number formats
    print("Fixing the number formats...")
    for i in range(3, sheet.max_column + 1):
        col = get_column_letter(i)
        for i in range(5, maxrow + 1):
            if col != "E":
                sheet[f"{col}{i}"].number_format = numbers.BUILTIN_FORMATS[3]
            else:
                sheet[f"{col}{i}"].number_format = numbers.BUILTIN_FORMATS[4]
        
    #fixes the borders
    print("Fixing borders...")
    for i in range(1, sheet.max_column + 1):
        col = get_column_letter(i)
        for i in range(5, maxrow + 1):
            sheet[f"{col}{i}"].border = border
    #adds filter buttons to columns
    sheet.auto_filter.ref = f"A4:{get_column_letter(sheet.max_column)}{maxrow}"

    print("Saving the workbook...\n")
    #saves the workbook
    workbook.save(default_base_file_path + fixed_file_name)

    end = process_time()
    excel_manip_time = "Time to manipulate excel: " + str(end - start) + "\n"

    #prints all the times to do certain sections of code. Commented out unless needed.
    print(fix_date_as_datetime_time)
    print(group_by_table_time)
    print(multiindex_table_time)
    print(fix_months_column_headers_time)
    print(make_total_sales_and_lbs_time)
    print(merge_levels_time)
    print(ot_export_fixed_time)
    print(excel_manip_time)

start2 = process_time()   
input_monthly_ytd_summary()
end2 = process_time()
#prints total time to run input_monthly_ytd_summary(). 
print("Total time: " + str(end2 - start2))


